# Author @Bhanu Surendra Deepala
# Automate sending Text & Media via Whatsapp using Python-selenium

'''
prerequisites
step1:Download & install PyCharm ( use below link )
https://www.jetbrains.com/pycharm/download/#section=windows
step2:Download chromedriver( use below link)
https://chromedriver.chromium.org/downloads
(download compatible version based on your system installed chrome version)
step3:Install Selenium using Terminal in PyCharm(after creating project)
use below command & press Enter:
    pip install selenium
step4:Install openpyxl to access data from excel
use below command & press Enter:
    pip install openpyxl

'''

# Importing required libraries
from selenium import webdriver
import openpyxl
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import easygui
from selenium.common.exceptions import NoSuchElementException

list_Ph=[] # Creating empty list to store all ph no from excel
counter=0
invalid_PhNo_List=[]
messagesuccess_List=[]
success_flag= False


# Bhanu Macbook
file1path = '/Users/bhanusurendradeepala/Downloads/KORM_MockTest.jpeg'
file2path = '/Users/bhanusurendradeepala/Downloads/KORM_CollegeCode.jpeg'

# Surekha Windows machine
# file1path = 'C:\\Users\\my\\Downloads\\KORM_MockTest.jpeg'
# file2path = 'C:\\Users\\my\\Downloads\\KORM_CollegeCode.jpeg'


textLink='https://docs.google.com/forms/d/e/1FAIpQLScqtXuFx1hDHlzCjk-CO1Vz5e4l6tjxjUcWXGJrB_zTd4x74w/viewform'
textMsg_updated='Dear Sir/Madam, Our college is conducting online mock EAPCET, please do register, you can also forward the above link to your friends and relatives, we really appreciate it. Please note the last date for Registration is 4th August, 2021 before 5:00 pm & exam will be conducted on 7th August, 2021. Top 3 scoring students will get FREE EDUCATION and ACCOMMODATION for 4 years, 4th and 5th place students will get FREE ACCOMMODATION for 4 years, kindly utilize the opportunity - Thank you.'
textMsg= '''
Dear Sir/Madam, Our college is conducting online mock EAPCET, please do register, you can also forward the above link to your friends and relatives, we really appreciate it.
Please note the last date for Registration is 4th August, 2021 before 5:00 pm & exam will be conducted on 7th August, 2021.
Top 3 scoring students will get FREE EDUCATION and ACCOMMODATION for 4 years, kindly utilize the opportunity - Thank you.'''

# Bhanu Macbook
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-automation'])
driver = webdriver.Chrome(executable_path='/Users/bhanusurendradeepala/Downloads/chromedriver_V92',options=options)
driver.implicitly_wait(30)

# Surekha Windows
# options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-automation'])
# driver = webdriver.Chrome(executable_path='C:\\Users\\my\\Downloads\\chromedriver',options=options)
# driver.implicitly_wait(30)


# Function to read phone numbers from excel
def read_data(path):

    datalist = [] # create empty list

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active

    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column

    # Loop will print all values in each row
    for i in range(2, max_row + 1):
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            value=str(cell_obj.value)
            if cell_obj.value is not None and value.__len__() > 10:
                datalist.append(cell_obj.value)
    return datalist


# Function to identify if element exists
def check_element_exists(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


# Function to initiate chat with required phone number
def initiateChat():

    chat_button = driver.find_element_by_xpath(
                '//*[@id="action-button"]')
    chat_button.click()

    web_button = driver.find_element_by_xpath(
                '//*[@id="fallback_block"]/div/div/a')
    web_button.click()

    invalid_alert_xpath='//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div[1]'

    if check_element_exists(invalid_alert_xpath):
        print('Please wait 20 sec, program will resume')
        return False
    else:
        return True


# Function to finally display successfully sent phone numbers
def displaySuccessfulMsg(counter):
    if counter>0:
        print("Files have been successfully sent to total : {} members, please see below".format(len(messagesuccess_List)))
    else:
        print("No Messages are sent out due to failure, please try again")
    for j in range(0,len(messagesuccess_List)):
        print('+{}'.format(messagesuccess_List[j]))


# Function to create & update output.txt file
def update_outputfile(counter,fileName,final_list,flag):
    if counter>0:
        with open(fileName, 'w') as f:
            if flag:
                f.write('Files have been successfully sent to total : {} members, please see below \n'.format(len(final_list)))
            else:
                f.write('Whatsapp was not installed for a total of : {} numbers, please see below  \n'.format(len(final_list)))

            for k in range(0,len(final_list)):
                ph_ext='+'+str(final_list[k])
                f.write(ph_ext)
                f.write('\n')
    else:
        with open(fileName, 'w') as f:
            if flag:
                f.write('No Messages are sent out due to failure, please try again')
            else:
                f.write('Whatsapp is installed for all numbers, no failures')


# Function to send Text Messages in Whatsapp
def sendTextMessage(text):
    input_box=driver.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[1]/div/div[2]')
    input_box.send_keys(text)
    input_box.click()
    sleep(8)
    button_send=driver.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]/button')
    button_send.click()
    sleep(12)
    return True


# Function to send media file and link in whatsapp
def sendFiles(filepath):
    try:
        # wait 20 seconds before looking for element
        element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//div[@title = "Attach"]'))
        )
    except:
        pass

    attachment_box = driver.find_element_by_xpath('//div[@title = "Attach"]')
    attachment_box.click()

    image_box = driver.find_element_by_xpath(
        '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')
    image_box.send_keys(filepath)

    # text_box=driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div[1]/span/div/div[2]/div/div[3]/div[1]/div[2]')
    # text_box.send_keys(textLink)

    send_button = driver.find_element_by_xpath(
        '//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/span/div/div/span')
    send_button.click()
    sleep(12)
    return True


# Program execution starts here
try:
    list_Ph = read_data('StudentsData.xlsx')
    print("You have 20 seconds to scan QR code!!")
    # easygui.msgbox('''Are you Ready to scan QR code!!
    #         You have only 20 seconds to scan QR code.
    #         Please keep your phone ready & Open Whatsapp
    #                 Goto Settings->Linked Devices ->Link a Device
    #                 Once you are READY, Click OK to Proceed!!!!''', title="SCAN QR CODE")

    # Loop through all phone numbers in the list
    for i in range(0,len(list_Ph)):
        sleep(4)
        Ph=list_Ph[i]
        Ph=str(Ph)
        # create custom url for sending whatsapp message
        url = 'https://api.whatsapp.com/send/?phone='+Ph
        # url='https://wa.me/'+Ph
        # printing to whom are we sending message for logging purpose
        print('Attempting to send Whatspp message to : +'+Ph)

        driver.get(url) # open browser with custom url

        # Calling functions in the sequence that is required
        flag = initiateChat()
        success_flag=False
        if flag:
            success_flag=sendTextMessage(textLink)
            success_flag=sendTextMessage(textMsg)
            success_flag=sendFiles(file1path)
            success_flag=sendFiles(file2path)
            # counter+=1
            if success_flag:
                print("Successfully sent files")
                messagesuccess_List.append(Ph)
                counter=len((messagesuccess_List))
                sleep(5)
        else:
            print("Whatsapp was not installed for :"+Ph)
            invalid_PhNo_List.append(Ph)
            success_flag=False
            sleep(20)


except Exception as e:
    print("Some Exception occured, please review the log/console and try again")
    # print("Actual Exception Details :")
    print(e)
    # displaySuccessfulMsg(counter)

finally:
    driver.close()
    driver.quit()
    displaySuccessfulMsg(counter)
    update_outputfile(counter,'output_successful.txt',messagesuccess_List,True)
    update_outputfile(counter, 'output_invalid_phno.txt', invalid_PhNo_List, False)
